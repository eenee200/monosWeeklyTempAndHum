import os
import requests
from datetime import datetime, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re as _re

# ── Configuration ──────────────────────────────────────────────────────────────
RECEIVER_EMAILS = os.environ.get('RECIPIENT_EMAIL')

CONFIG = {
    'GPS_API_KEY': os.environ.get('API_KEY'),
    'VEHICLES': {
        '868373075408486': '3922УБЯ',
        '350544501468303': '5034УКН',
        '863719068034074': '5035УКН',
        '350317174707566': '5036УКН'
    },
    'SENDER_EMAIL':   os.environ.get('GMAIL_EMAIL'),
    'SENDER_PASSWORD': os.environ.get('GMAIL_PASSWORD'),
    'RECEIVER_EMAILS': RECEIVER_EMAILS.split(','),
}

IO_TEMP      = 'io10800'
IO_TEMP2     = 'io25'
IO_HUMIDITY  = 'io10804'
IO_HUMIDITY2 = 'io86'

FONT_NAME   = "Arial"
C_HEADER_BG = "1F4E79"
C_HEADER_FG = "FFFFFF"
C_SUBHDR_BG = "2E75B6"
C_SUBHDR_FG = "FFFFFF"
C_LABEL_BG  = "D6E4F0"
C_ALT_ROW   = "EBF3FA"
C_RED_BG    = "FFE0E0"
C_RED_FG    = "C00000"
C_BORDER    = "8EA9C1"

# Each car block: 3 data cols + 1 gap col
BLOCK_COLS = 4
DATA_COLS  = 3

thin = Side(style="thin", color=C_BORDER)

def _border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, color=color, name=FONT_NAME)

def _align(h="center", v="center"):
    return Alignment(horizontal=h, vertical=v)

def _wc(ws, row, col, value, bold=False, size=10, fg="000000",
        bg=None, halign="center", border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, size=size, color=fg)
    cell.alignment = _align(h=halign)
    if bg:
        cell.fill = _fill(bg)
    if border:
        cell.border = _border()
    return cell


# ═══════════════════════════════════════════════════════════════════════════════
#  FETCH & PARSE
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_vehicle_data(device_id, start_str, end_str, api_key):
    url = (f"https://fms2.gpsbox.mn/api/api.php?api=user&key={api_key}"
           f"&cmd=OBJECT_GET_MESSAGES,{device_id},{start_str},{end_str},0.01")
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    return r.json()


def parse_api_response(json_data):
    last_valid_temp = None
    last_valid_hum  = None
    daily_first = {}
    daily_last  = {}

    for entry in json_data:
        try:
            ts    = datetime.strptime(entry[0], '%Y-%m-%d %H:%M:%S') + timedelta(hours=8)
            speed = float(entry[5]) if len(entry) > 5 else 0.0
        except Exception:
            continue
        if speed > 0:
            d = ts.date()
            if d not in daily_first:
                daily_first[d] = ts
            daily_last[d] = ts

    daily_temp = defaultdict(list)
    daily_hum  = defaultdict(list)

    for entry in json_data:
        try:
            ts = datetime.strptime(entry[0], '%Y-%m-%d %H:%M:%S') + timedelta(hours=8)
            io = entry[6] if len(entry) > 6 and isinstance(entry[6], dict) else {}
        except Exception:
            continue

        d         = ts.date()
        in_window = (d in daily_first and daily_first[d] <= ts <= daily_last[d])

        temp_key = IO_TEMP if IO_TEMP in io else (IO_TEMP2 if IO_TEMP2 in io else None)
        if temp_key:
            raw = float(io[temp_key])
            if raw == 250:
                temp = last_valid_temp
            else:
                temp = raw / 100.0
                last_valid_temp = temp
            if temp is not None and in_window:
                daily_temp[d].append({'timestamp': ts, 'temperature': temp})

        hum_key = IO_HUMIDITY if IO_HUMIDITY in io else (IO_HUMIDITY2 if IO_HUMIDITY2 in io else None)
        if hum_key:
            raw = float(io[hum_key])
            hum = last_valid_hum if raw == 250 else raw
            if raw != 250:
                last_valid_hum = raw
            if hum is not None and in_window:
                daily_hum[d].append({'timestamp': ts, 'humidity': hum})

    return daily_temp, daily_hum


# ═══════════════════════════════════════════════════════════════════════════════
#  STATS
# ═══════════════════════════════════════════════════════════════════════════════

def calc_temp_stats(td):
    if not td:
        return None
    temps = [r['temperature'] for r in td]
    dh, R = 83144.0, 8.314
    tk    = [t + 273.15 for t in temps]
    mkt   = (dh / R) / (sum(dh / (R * t) for t in tk) / len(tk)) - 273.15
    return {'highest': max(temps), 'lowest': min(temps),
            'average': sum(temps)/len(temps), 'mkt': mkt,
            'count': len(temps),
            'start_time': td[0]['timestamp'], 'stop_time': td[-1]['timestamp']}

def calc_hum_stats(hd):
    if not hd:
        return None
    h = [r['humidity'] for r in hd]
    return {'highest': max(h), 'lowest': min(h),
            'average': sum(h)/len(h), 'count': len(h)}

def fmt_elapsed(t1, t2):
    s = int((t2-t1).total_seconds())
    return f"{s//3600}h {(s%3600)//60}m"


# ═══════════════════════════════════════════════════════════════════════════════
#  10-MINUTE RESAMPLING  (for bottom data table only)
# ═══════════════════════════════════════════════════════════════════════════════

def resample_10min(temp_data, humidity_data):
    """
    Bucket raw readings into 10-minute windows.
    Each bucket is labelled by its window start (floor to 10 min).
    Returns a list of dicts: {timestamp, temperature, humidity}
    temperature and humidity are averages of all readings in that window.
    """
    temp_buckets = defaultdict(list)
    hum_buckets  = defaultdict(list)

    for r in temp_data:
        ts = r['timestamp']
        # floor to nearest 10-minute mark
        bucket = ts.replace(minute=(ts.minute // 10) * 10, second=0, microsecond=0)
        temp_buckets[bucket].append(r['temperature'])

    for r in humidity_data:
        ts = r['timestamp']
        bucket = ts.replace(minute=(ts.minute // 10) * 10, second=0, microsecond=0)
        hum_buckets[bucket].append(r['humidity'])

    all_buckets = sorted(set(temp_buckets) | set(hum_buckets))

    result = []
    for bucket in all_buckets:
        temps = temp_buckets.get(bucket)
        hums  = hum_buckets.get(bucket)
        result.append({
            'timestamp':   bucket,
            'temperature': sum(temps) / len(temps) if temps else None,
            'humidity':    sum(hums)  / len(hums)  if hums  else None,
        })

    return result

# ═══════════════════════════════════════════════════════════════════════════════
#  DAILY SHEET — all cars left → right
# ═══════════════════════════════════════════════════════════════════════════════

def build_daily_sheet(wb, date, plates, vehicle_data):
    date_label = date.strftime('%Y-%m-%d')
    ws = wb.create_sheet(title=date_label)
    ws.sheet_view.showGridLines = False

    n_cars = len(plates)

    # Column widths
    for ci in range(n_cars):
        base = ci * BLOCK_COLS + 1
        ws.column_dimensions[get_column_letter(base)    ].width = 10  # time
        ws.column_dimensions[get_column_letter(base + 1)].width = 7   # °C
        ws.column_dimensions[get_column_letter(base + 2)].width = 7   # %RH
        ws.column_dimensions[get_column_letter(base + 3)].width = 2   # gap

    # Collect data
    car_td    = []
    car_hd    = []
    car_s     = []
    car_hs    = []
    car_hdict = []
    for plate in plates:
        dt, dh = vehicle_data[plate]
        td = dt.get(date, [])
        hd = dh.get(date, [])
        car_td.append(td)
        car_hd.append(hd)
        car_s.append(calc_temp_stats(td))
        car_hs.append(calc_hum_stats(hd))
        car_hdict.append({r['timestamp']: r['humidity'] for r in hd})

    # ── Row 1: sheet title ─────────────────────────────────────────────────────
    row = 1
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=1, value=f"Дата тайлан — {date_label}")
    c.font = _font(bold=True, size=13, color=C_HEADER_BG)
    c.alignment = _align(h="left")
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_cars * BLOCK_COLS)
    row += 1

    # ── Row 2: car name banners ────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    for ci, plate in enumerate(plates):
        base = ci * BLOCK_COLS + 1
        c = ws.cell(row=row, column=base, value=f"🚛 {plate}")
        c.font = _font(bold=True, size=11, color=C_HEADER_FG)
        c.fill = _fill(C_HEADER_BG)
        c.alignment = _align(h="left")
        c.border = _border()
        for dc in range(1, DATA_COLS):
            fc = ws.cell(row=row, column=base+dc)
            fc.fill = _fill(C_HEADER_BG); fc.border = _border()
        ws.merge_cells(start_row=row, start_column=base,
                       end_row=row, end_column=base + DATA_COLS - 1)
    row += 1

    # ── Rows 3-8: stats panel (6 rows) ────────────────────────────────────────
    STAT_ROWS = 6
    for sr in range(STAT_ROWS):
        ws.row_dimensions[row + sr].height = 15

    stat_defs = [
        lambda s, hs: ("Эхэлсэн",  s['start_time'].strftime('%H:%M:%S') if s else "-", ""),
        lambda s, hs: ("Эцсийн",   s['stop_time'].strftime('%H:%M:%S')  if s else "-", ""),
        lambda s, hs: ("Нийт цаг", fmt_elapsed(s['start_time'], s['stop_time']) if s else "-", ""),
        lambda s, hs: ("↑ / ↓ °C", f"{s['highest']:.1f}" if s else "-",
                                    f"{s['lowest']:.1f}"  if s else "-"),
        lambda s, hs: ("Дундаж / MKT", f"{s['average']:.1f}" if s else "-",
                                    f"{s['mkt']:.1f}"     if s else "-"),
        lambda s, hs: ("%RH ↑/Дундаж", f"{hs['highest']:.1f}" if hs else "-",
                                    f"{hs['average']:.1f}"  if hs else "-"),
    ]

    for sr, fn in enumerate(stat_defs):
        for ci in range(n_cars):
            base = ci * BLOCK_COLS + 1
            label, v1, v2 = fn(car_s[ci], car_hs[ci])
            _wc(ws, row+sr, base,   label, bold=True,  bg=C_LABEL_BG, halign="left", size=9)
            _wc(ws, row+sr, base+1, v1,    bold=False, bg=C_LABEL_BG, size=9)
            _wc(ws, row+sr, base+2, v2,    bold=False, bg=C_LABEL_BG, size=9)
    row += STAT_ROWS

    # ── Data column headers ────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 15
    for ci in range(n_cars):
        base = ci * BLOCK_COLS + 1
        for dc, lbl in enumerate(["Цаг", "°C", "%RH"]):
            _wc(ws, row, base+dc, lbl, bold=True, bg=C_SUBHDR_BG, fg=C_SUBHDR_FG, size=9)
    row += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    # ── Precompute 10-min resampled data per car ───────────────────────────────
    car_table = []
    for ci in range(n_cars):
        car_table.append(resample_10min(car_td[ci], car_hd[ci]))

    # ── Data rows ─────────────────────────────────────────────────────────────
    max_rows = max((len(t) for t in car_table), default=0)

    for i in range(max_rows):
        alt = (i % 2 == 1)
        ws.row_dimensions[row+i].height = 13

        for ci in range(n_cars):
            base       = ci * BLOCK_COLS + 1
            table_data = car_table[ci]
            ebg        = C_ALT_ROW if alt else "FFFFFF"

            if i >= len(table_data):
                for dc in range(DATA_COLS):
                    c = ws.cell(row=row+i, column=base+dc, value="")
                    c.border = _border(); c.fill = _fill(ebg)
                continue

            reading = table_data[i]
            t  = reading['temperature']
            hv = reading['humidity']
            oor = t is not None and (t <= -9 or t >= 15)
            rbg = C_RED_BG if oor else ebg
            tfg = C_RED_FG if oor else "000000"

            ts_c = ws.cell(row=row+i, column=base,
                           value=reading['timestamp'].strftime('%H:%M'))
            ts_c.font = _font(size=8); ts_c.alignment = _align(h="center")
            ts_c.border = _border(); ts_c.fill = _fill(rbg)

            t_val = round(t, 1) if t is not None else "-"
            t_c = ws.cell(row=row+i, column=base+1, value=t_val)
            t_c.font = _font(size=8, color=tfg, bold=oor)
            t_c.alignment = _align(h="center"); t_c.border = _border()
            t_c.fill = _fill(rbg)
            if isinstance(t_val, float):
                t_c.number_format = '0.0'

            hval = round(hv, 1) if hv is not None else "-"
            h_c  = ws.cell(row=row+i, column=base+2, value=hval)
            h_c.font = _font(size=8); h_c.alignment = _align(h="center")
            h_c.border = _border(); h_c.fill = _fill(rbg)
            if isinstance(hval, float):
                h_c.number_format = '0.0'

    row += max_rows

    # ── OOR section ───────────────────────────────────────────────────────────
    oor_lists = [[r for r in td if r['temperature'] <= -9 or r['temperature'] >= 15]
                 for td in car_td]

    if any(oor_lists):
        row += 1
        ws.row_dimensions[row].height = 17

        for ci in range(n_cars):
            base = ci * BLOCK_COLS + 1
            oor  = oor_lists[ci]
            lbl  = f"⚠️ OOR ({len(oor)})" if oor else "✅ OOR байхгүй"
            bg   = "C00000" if oor else "2E7D32"
            c = ws.cell(row=row, column=base, value=lbl)
            c.font = _font(bold=True, size=9, color="FFFFFF")
            c.fill = _fill(bg); c.alignment = _align(h="center"); c.border = _border()
            for dc in range(1, DATA_COLS):
                fc = ws.cell(row=row, column=base+dc)
                fc.fill = _fill(bg); fc.border = _border()
            ws.merge_cells(start_row=row, start_column=base,
                           end_row=row, end_column=base+DATA_COLS-1)
        row += 1

        ws.row_dimensions[row].height = 14
        for ci in range(n_cars):
            base = ci * BLOCK_COLS + 1
            for dc, lbl in enumerate(["Цаг", "°C", "%RH"]):
                _wc(ws, row, base+dc, lbl, bold=True, bg="FF4444", fg="FFFFFF", size=9)
        row += 1

        max_oor = max((len(o) for o in oor_lists), default=0)
        for i in range(max_oor):
            alt = (i % 2 == 1)
            bg  = "FFD0D0" if alt else C_RED_BG
            ws.row_dimensions[row+i].height = 13

            for ci in range(n_cars):
                base  = ci * BLOCK_COLS + 1
                oor   = oor_lists[ci]
                hdict = car_hdict[ci]

                if i >= len(oor):
                    for dc in range(DATA_COLS):
                        c = ws.cell(row=row+i, column=base+dc, value="")
                        c.border = _border(); c.fill = _fill(bg)
                    continue

                rec = oor[i]
                t   = rec['temperature']
                hv  = hdict.get(rec['timestamp'])

                ts_c = ws.cell(row=row+i, column=base,
                               value=rec['timestamp'].strftime('%H:%M:%S'))
                ts_c.font = _font(size=8); ts_c.alignment = _align(h="center")
                ts_c.border = _border(); ts_c.fill = _fill(bg)

                t_c = ws.cell(row=row+i, column=base+1, value=round(t, 1))
                t_c.font = _font(size=8, color=C_RED_FG, bold=True)
                t_c.alignment = _align(h="center"); t_c.border = _border()
                t_c.fill = _fill(bg); t_c.number_format = '0.0'

                hval = round(hv, 1) if hv is not None else "-"
                h_c  = ws.cell(row=row+i, column=base+2, value=hval)
                h_c.font = _font(size=8); h_c.alignment = _align(h="center")
                h_c.border = _border(); h_c.fill = _fill(bg)
                if isinstance(hval, float):
                    h_c.number_format = '0.0'

    ws.freeze_panes = "A3"
    print(f"  Sheet: {date_label}")


# ═══════════════════════════════════════════════════════════════════════════════
#  SUMMARY SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_summary_sheet(wb, all_dates, plates, vehicle_data):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    for i, w in enumerate([12,18,12,12,12,12,12,12,12,10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws.row_dimensions[row].height = 24
    c = ws.cell(row=row, column=1, value="Дата тайлан — Нэгтгэл")
    c.font = _font(bold=True, size=14, color=C_HEADER_BG)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 2

    for col, h in enumerate(["Огноо","Машин","Эхэлсэн","Эцсийн",
                              "Өндөр°C","Бага°C","Дундаж°C","MKT°C",
                              "Өндөр%RH","Тоо"], 1):
        _wc(ws, row, col, h, bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, size=9)
    row += 1

    idx = 0
    for date in all_dates:
        for plate in plates:
            dt, dh = vehicle_data[plate]
            s  = calc_temp_stats(dt.get(date, []))
            hs = calc_hum_stats(dh.get(date, []))
            bg = C_ALT_ROW if idx % 2 else "FFFFFF"
            if s:
                vals = [date.strftime('%Y-%m-%d'), plate,
                        s['start_time'].strftime('%H:%M'),
                        s['stop_time'].strftime('%H:%M'),
                        round(s['highest'],2), round(s['lowest'],2),
                        round(s['average'],2), round(s['mkt'],2),
                        round(hs['highest'],2) if hs else "-", s['count']]
            else:
                vals = [date.strftime('%Y-%m-%d'), plate,
                        "-","-","-","-","-","-","-", 0]
            for col, v in enumerate(vals, 1):
                _wc(ws, row, col, v, bg=bg, size=9)
            row += 1; idx += 1

    ws.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════════════════
#  EMAIL
# ═══════════════════════════════════════════════════════════════════════════════

def send_email_with_attachment(sender_email, sender_password, receiver_emails,
                               subject, message, attachment_path):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    from email.mime.text import MIMEText

    if isinstance(receiver_emails, str):
        receiver_emails = [receiver_emails]
    receiver_emails = [e.strip() for e in receiver_emails if e.strip()]
    if not receiver_emails:
        print("No valid receiver emails"); return False
    try:
        msg = MIMEMultipart()
        msg['From'] = sender_email; msg['To'] = ', '.join(receiver_emails)
        msg['Subject'] = subject; msg.attach(MIMEText(message, 'plain'))
        with open(attachment_path, 'rb') as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(attachment_path))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(attachment_path)}"'
            msg.attach(part)
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls(); server.login(sender_email, sender_password)
            server.send_message(msg)
        print(f"Email sent to: {', '.join(receiver_emails)}")
        return True
    except Exception as e:
        print(f"Email error: {e}"); return False


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    today       = datetime.now()
    
    last_monday   = (today - timedelta(days=today.weekday() + 7)).replace(
        hour=0, minute=0, second=0, microsecond=0)
    last_sunday = last_monday + timedelta(days=7)

    start_str = (last_monday   - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M')
    end_str   = (last_sunday - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M')
    print(f"Fetching: {start_str} → {end_str}\n")

    plates       = list(CONFIG['VEHICLES'].values())
    vehicle_data = {}

    for device_id, plate in CONFIG['VEHICLES'].items():
        print(f"Fetching {plate}…")
        try:
            raw = fetch_vehicle_data(device_id, start_str, end_str, CONFIG['GPS_API_KEY'])
            vehicle_data[plate] = parse_api_response(raw)
        except Exception as e:
            print(f"  ERROR: {e}")
            vehicle_data[plate] = (defaultdict(list), defaultdict(list))

    all_dates = sorted(set(
        d
        for dt, dh in vehicle_data.values()
        for d in (set(dt) | set(dh))
    ))

    if not all_dates:
        print("No data."); return

    os.makedirs('reports', exist_ok=True)
    output_file = 'reports/sensor_report.xlsx'

    wb = Workbook()
    build_summary_sheet(wb, all_dates, plates, vehicle_data)

    for date in all_dates:
        build_daily_sheet(wb, date, plates, vehicle_data)

    wb.save(output_file)
    print(f"\nSaved: {output_file}")

    send_email_with_attachment(
        CONFIG['SENDER_EMAIL'], CONFIG['SENDER_PASSWORD'], RECEIVER_EMAILS,
        f"Sensor Report — {last_sunday.date()}",
        f"Period: {last_monday.date()} → {last_sunday.date()}",
        output_file,
    )

if __name__ == "__main__":
    main()
